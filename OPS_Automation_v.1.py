# =============================================================================
# 
# Author    : Venna Naga Durgaprasad/IBM
#             Venna.Naga.Durgaprsad@ibm.com
# Date      : 28-03-2024
# Version   : 0.1
# =============================================================================
# Purpose   : Automation of OPS - MS, Empower, RS-GROD      
# =============================================================================

# ------------------------ < IMPORTING PACKAGE > ------------------------------
import os
import sys
from datetime import datetime as dt,timedelta
import time
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl.styles import PatternFill, Font, Color, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ------------------- < CREATE/OPEN the EXCEL Report > ------------------------
app_path=os.path.dirname(sys.argv[0])                               # Get current application path
#print(app_path)
#input_file=input('Please enter the Roster file name: ')
#output_file=input('Please  enter the output file name: ')
#wb=Workbook(f'{input_file}')

input_dt_ot=''
def Welcome():
    def delay_print(s):
        for c in s:
            sys.stdout.write(c)
            sys.stdout.flush()
            time.sleep(0.10)        
    welcome_note="         >>> Welcome to OPS Automation Tool  <<<         "
    delay_print(f"\n\033[20;1;4;5;41m{welcome_note}\033[0m")
    
    print('\n\n\n\033[32;1mCurrent Folder: \033[0m'+f'\033[33;1m{app_path}\033[0m')
    print("\n\n\033[34;1m\n\nPlease prepare Roster header as per requirement: \n\033[0m")
    roster_note=["Account Id","LOB","EMP ID", "XID", "Name", "Team", "Site", "Status", "Work item", "Billable / Bench", "Type of Emp", "PIR Location", "Billing Start Date", "Billing End Date", "End Date"]
    for note in roster_note:
        sys.stdout.write(f"\033[97;1;1m{note}\033[0m\n")
        sys.stdout.flush()
        time.sleep(0.10) 
    time.sleep(1)
    print("\n\n>>> Please make sure Enter details correctly \n")
    time.sleep(1)

def date_checker(input_dt):
    try: 
        global input_dt_ot
        input_dt_ot=input_dt
        global check_dt
        check_dt=dt.strptime(input_dt,"%d-%m-%Y")
        global output_dt
        global check_wk_dt
        output_dt=check_dt.strftime("%Y-%m-%d")
        check_wk_dt=check_dt.weekday() in [4] # if is weekend or not
        if check_wk_dt==False:
            print('\033[31;1mInput date is not weekend date\033[0m')
            new_input_dt=input('Enter Weekending Date (DD-MM-YYYY): ')
            date_checker(new_input_dt)
    except ValueError or NameError:
        print('\033[31;1mInput date is Invalid date format\033[0m')
        new_input_dt=input('Enter Weekending Date (DD-MM-YYYY): ')
        date_checker(new_input_dt)

def load_workbook(filename):
    try:
        return openpyxl.load_workbook(filename)
    except FileNotFoundError as file:
        print(f"\033[31;1m{filename} does not exist. Please check the file name before running the code.\033[0m")
        new_filename=input('Enter correct file name: ')
        return load_workbook(f'{new_filename}.xlsx')

def alt_file(filename):
    if filename=='skip.xls':
        print(f'{filename} is skipped')
    else:
        try:
            return pd.read_excel(filename)
        except FileNotFoundError as file:
            print(f"\033[31;1m{filename} does not exist. Please check the file name before running the code. file format in .xls\033[0m")
            new_filename=input('Enter correct file name: ')
            return alt_file(f'{new_filename}.xls')

def output_report():
    global wb21
    global wbs21
    global wbs22
    global wbs23

    wb21=openpyxl.Workbook() # Create a new workbook
    wbs21=wb21.active          # Create a sheet in the workbook
    # <<<<<<<<<<<<<<<<<<<<<<<< Header & Creating Sheets >>>>>>>>>>>>>>>>>>>>>>>>

    # ------------------------- < RS-MS sheet  > -------------------------------
    wbs21.title="RS-MS"
    wbs21.sheet_properties.tabColor= "36A832"       
    wbs21.append(["Account Id", "EMP ID", "XID", "Name", "Team", "Site", "Status", "Work item", "Billable / Bench", "Type of Emp", "PIR Location", "Billing Start Date", "Billing End Date", "End Date","ILC","Remarks"])
    '''
    wbs21['A1'].value="Account Id"
    wbs21['B1'].value="EMP ID"
    wbs21['C1'].value="XID"
    wbs21['D1'].value="Name"
    wbs21['E1'].value="Team"
    wbs21['F1'].value="Site"
    wbs21['G1'].value="Status"
    wbs21['H1'].value="Work item"
    wbs21['I1'].value="Billable / Bench"
    wbs21['J1'].value="Type of Emp"
    wbs21['K1'].value="PIR Location"
    wbs21['L1'].value="Billing Start Date"
    wbs21['M1'].value="Billing End"
    wbs21['N1'].value="End Date"

    '''

    # ------------------------- < EMPOWER sheet > -------------------------------
    wb21.create_sheet('EMPOWER')
    wbs22=wb21["EMPOWER"]
    wbs22.sheet_properties.tabColor= "FF0000"  
    wbs22.append(["Account Id", "EMP ID", "XID", "Name", "Team", "Site", "Status", "Work item", "Billable / Bench", "Type of Emp", "PIR Location", "Billing Start Date", "Billing End Date", "End Date","ILC","Tempo","Match","Remarks"])
    '''
    wbs22['A1'].value="Account Id"
    wbs22['B1'].value="EMP ID"
    wbs22['C1'].value="XID"
    wbs22['D1'].value="Name"
    wbs22['E1'].value="Team"
    wbs22['F1'].value="Site"
    wbs22['G1'].value="Status"
    wbs22['H1'].value="Work item"
    wbs22['I1'].value="Billable / Bench"
    wbs22['J1'].value="Type of Emp"
    wbs22['K1'].value="PIR Location"
    wbs22['L1'].value="Billing Start Date"
    wbs22['M1'].value="Billing End Date"
    wbs22['N1'].value="End Date"
    '''

    # ------------------------- < RS-GROD sheet > -------------------------------
    wb21.create_sheet('RS-GROD')
    wbs23=wb21["RS-GROD"]
    wbs23.sheet_properties.tabColor= "0037FF"  
    wbs23.append(["Account Id", "EMP ID", "XID", "Name", "Team", "Site", "Status", "Work item", "Billable / Bench", "Type of Emp", "PIR Location", "Billing Start Date", "Billing End Date", "End Date", "ACC ID", "ACC ID Match", "Work Item", "WI Match", "Billable", "CNB", "EPIC", "EPIC Match", "Remarks"])
    '''
    wbs23['A1'].value="Account Id"
    wbs23['B1'].value="EMP ID"
    wbs23['C1'].value="XID"
    wbs23['D1'].value="Name"
    wbs23['E1'].value="Team"
    wbs23['F1'].value="Site"
    wbs23['G1'].value="Status"
    wbs23['H1'].value="Work item"
    wbs23['I1'].value="Billable / Bench"
    wbs23['J1'].value="Type of Emp"
    wbs23['K1'].value="PIR Location"
    wbs23['L1'].value="Billing Start Date"
    wbs23['M1'].value="Billing End Date"
    wbs23['N1'].value="End Date"
    wbs23['O1'].value="ACC ID"
    wbs23['P1'].value="ACC ID Match"
    wbs23['Q1'].value="Work Iteam"
    wbs23['R1'].value="WI Match"
    wbs23['S1'].value="Billable"
    wbs23['T1'].value="CNB"
    wbs23['U1'].value="EPIC"
    wbs23['V1'].value="EPIC Match"
    wbs23['W1'].value="Remarks"
    '''
    # ---------------------------------------------------------------------------

class Employee:
    def __init__(self,emp_id,account_id,lob,xid,name, team, site, status, work_item , billable_bench ,type_of_emp, PIR_location=None, bill_start_dt=None,bill_end_dt=None, end_date=None):
        self.emp_id=emp_id
        self.account_id=account_id
        self.lob=lob
        self.xid=xid
        self.name=name
        self.team=team
        self.site=site
        self.status=status
        self.work_item=work_item
        self.billable_bench=billable_bench
        self.type_of_emp=type_of_emp
        self.PIR_location=PIR_location
        self.bill_start_dt=bill_start_dt
        self.bill_end_dt=bill_end_dt
        self.end_date=end_date

class PRIMA_MS():
    def __init__(self, p_emp_id, p_account_id, p_work_item,p_weekend_date,p_activity_code,p_total_hours=0):
        self.p_emp_id=p_emp_id
        self.p_account_id=p_account_id
        self.p_work_item=p_work_item
        self.p_weekend_date=p_weekend_date
        self.p_activity_code=p_activity_code
        self.p_total_hours=p_total_hours
    def update(self, new_weekend_date, new_activity_code, hours):
        if self.p_weekend_date==new_weekend_date and self.p_activity_code==new_activity_code:
            self.p_total_hours+=hours
            
class PRIMA_EMPOWER():
    def __init__(self, p_emp_id, p_account_id, p_work_item,p_weekend_date,p_activity_code,p_total_hours):
        self.p_emp_id=p_emp_id
        self.p_account_id=p_account_id
        self.p_work_item=p_work_item
        self.p_weekend_date=p_weekend_date
        self.p_activity_code=p_activity_code
        self.p_total_hours=p_total_hours
    def update(self, new_weekend_date, new_activity_code, hours):
        if self.p_weekend_date==new_weekend_date and self.p_activity_code==new_activity_code:
            self.p_total_hours+=hours

class PRIMA_GROD():
    def __init__(self, p_emp_id, p_account_id, p_work_item,p_weekend_date,p_activity_code,p_total_hours=0,cnb=0):
        self.p_emp_id=p_emp_id
        self.p_account_id=p_account_id
        self.p_work_item=p_work_item
        self.p_weekend_date=p_weekend_date
        self.p_activity_code=p_activity_code
        self.p_total_hours=p_total_hours
        self.cnb=cnb
    def update(self, new_weekend_date, new_activity_code, hours,cnb):
        if self.p_weekend_date==new_weekend_date and self.p_activity_code==new_activity_code:
            if new_activity_code=="GN0035":
                self.cnb+=cnb
            else:
                self.p_total_hours+=hours

def date_formatting(given_date):
    if given_date:
        given_dt=dt.strptime(str(given_date), '%Y-%m-%d %H:%M:%S')
        return given_dt.strftime('%m-%d-%Y')
'''unpaid_checker=True
def EPIC_checker(alt_xid,alt_dt):
    global unpaid_checker
 '''   

def roster_report():
    global emp_data
    emp_data=[]
    for row in range(wbs11.min_row+1,wbs11.max_row+1):
        emp_id=wbs11.cell(row=row, column=3).value.strip()
        if emp_id:
                emp_data.append(Employee(emp_id, wbs11.cell(row=row, column=1).value,wbs11.cell(row=row, column=2).value, wbs11.cell(row=row, column=4).value, wbs11.cell(row=row, column=5).value, wbs11.cell(row=row, column=6).value, wbs11.cell(row=row, column=7).value, wbs11.cell(row=row, column=8).value, wbs11.cell(row=row, column=9).value, wbs11.cell(row=row, column=10).value, wbs11.cell(row=row, column=11).value, wbs11.cell(row=row, column=12).value,wbs11.cell(row=row, column=13).value,wbs11.cell(row=row, column=14).value,wbs11.cell(row=row, column=15).value))

    counter1=2
    counter2=2
    counter3=2
    for emp in emp_data:
        if  emp.lob=='RS MS':
            wbs21[f'A{counter1}'].value=emp.account_id
            wbs21[f'B{counter1}'].value=emp.emp_id
            wbs21[f'C{counter1}'].value=emp.xid
            wbs21[f'D{counter1}'].value=emp.name
            wbs21[f'E{counter1}'].value=emp.team
            wbs21[f'F{counter1}'].value=emp.site
            wbs21[f'G{counter1}'].value=emp.status
            wbs21[f'H{counter1}'].value=emp.work_item
            wbs21[f'I{counter1}'].value=emp.billable_bench
            wbs21[f'J{counter1}'].value=emp.type_of_emp
            wbs21[f'K{counter1}'].value=emp.PIR_location
            wbs21[f'L{counter1}'].value=date_formatting(emp.bill_start_dt)
            wbs21[f'M{counter1}'].value=date_formatting(emp.bill_end_dt)
            wbs21[f'N{counter1}'].value=date_formatting(emp.end_date)
            counter1+=1

        elif emp.lob=="Empower":
            wbs22[f'A{counter2}'].value=emp.account_id
            wbs22[f'B{counter2}'].value=emp.emp_id
            wbs22[f'C{counter2}'].value=emp.xid
            wbs22[f'D{counter2}'].value=emp.name
            wbs22[f'E{counter2}'].value=emp.team
            wbs22[f'F{counter2}'].value=emp.site
            wbs22[f'G{counter2}'].value=emp.status
            wbs22[f'H{counter2}'].value=emp.work_item
            wbs22[f'I{counter2}'].value=emp.billable_bench
            wbs22[f'J{counter2}'].value=emp.type_of_emp
            wbs22[f'K{counter2}'].value=emp.PIR_location
            wbs22[f'L{counter2}'].value=date_formatting(emp.bill_start_dt)
            wbs22[f'M{counter2}'].value=date_formatting(emp.bill_end_dt)
            wbs22[f'N{counter2}'].value=date_formatting(emp.end_date)
            counter2+=1

        elif emp.lob=='RS-GROD':
            wbs23[f'A{counter3}'].value=emp.account_id
            wbs23[f'B{counter3}'].value=emp.emp_id
            wbs23[f'C{counter3}'].value=emp.xid.upper()
            wbs23[f'D{counter3}'].value=emp.name
            wbs23[f'E{counter3}'].value=emp.team
            wbs23[f'F{counter3}'].value=emp.site
            wbs23[f'G{counter3}'].value=emp.status
            wbs23[f'H{counter3}'].value=emp.work_item
            wbs23[f'I{counter3}'].value=emp.billable_bench
            wbs23[f'J{counter3}'].value=emp.type_of_emp
            wbs23[f'K{counter3}'].value=emp.PIR_location
            wbs23[f'L{counter3}'].value=date_formatting(emp.bill_start_dt)
            wbs23[f'M{counter3}'].value=date_formatting(emp.bill_end_dt)
            wbs23[f'N{counter3}'].value=date_formatting(emp.end_date)
            counter3+=1

def prima_ms():
    emp_prima_data_ms=[]
    for row in range(5,wbs12.max_row+1):
        emp_id=wbs12.cell(row=row, column=20).value.strip()
        if emp_id:
            p_account_id=wbs12.cell(row=row, column=1).value.strip()
            p_work_item=wbs12.cell(row=row, column=3).value.strip()
            new_p_weekend_date=wbs12.cell(row=row, column=26).value
            new_p_activity_code=wbs12.cell(row=row, column=8).value.strip()
            new_p_total_hours=int(wbs12.cell(row=row, column=40).value)
            match_emp=None
            for check_emp in emp_prima_data_ms:
                if check_emp.p_emp_id==emp_id:
                    if check_emp.p_weekend_date==new_p_weekend_date:
                        match_emp=check_emp
                        break
            if match_emp:
                match_emp.update(new_p_weekend_date, new_p_activity_code, new_p_total_hours)
            else:
                emp_prima_data_ms.append(PRIMA_MS(emp_id,p_account_id, p_work_item,new_p_weekend_date,new_p_activity_code,new_p_total_hours))

    counter1=2
    check_emp=True
    for emp in emp_data:
        
        if  emp.lob=='RS MS':
            for prima_data in emp_prima_data_ms:
                if prima_data.p_emp_id==emp.emp_id and prima_data.p_weekend_date==output_dt:
                    wbs21[f'O{counter1}'].value=prima_data.p_total_hours
                    check_emp=False
                    if int(wbs21[f'O{counter1}'].value)>45:
                            wbs21[f'P{counter1}'].value="Billable  Hour is more than 45."
                    break       
            if check_emp==True:
                wbs21[f'O{counter1}'].value=0
                wbs21[f'P{counter1}'].value="Emp data not found in PRIMA Report" 
            check_emp=True
            counter1+=1
            
def prima_empower():
    global emp_prima_data_empower
    emp_prima_data_empower=[]
    for row in range(5,wbs13.max_row+1):

        emp_id=wbs13.cell(row=row, column=20).value.strip()
        if emp_id:
            p_account_id=wbs13.cell(row=row, column=1).value.strip()
            p_work_item=wbs13.cell(row=row, column=3).value.strip()
            new_p_weekend_date=wbs13.cell(row=row, column=26).value
            new_p_activity_code=wbs13.cell(row=row, column=8).value.strip()
            new_p_total_hours=int(wbs13.cell(row=row, column=40).value)
            match_emp=None
            for check_emp in emp_prima_data_empower:
                if check_emp.p_emp_id==emp_id:
                    if check_emp.p_weekend_date==new_p_weekend_date and check_emp.p_activity_code==new_p_activity_code:
                        match_emp=check_emp
                        break
            if match_emp:
                match_emp.update(new_p_weekend_date, new_p_activity_code, new_p_total_hours)
            else:
                emp_prima_data_empower.append(PRIMA_EMPOWER(emp_id,p_account_id, p_work_item,new_p_weekend_date,new_p_activity_code,new_p_total_hours))
    #for prima_data in emp_prima_data_empower:
     #   print(prima_data.p_weekend_date)
    check_emp=True
    counter2=2
    for emp in emp_data:
        if emp.lob=="Empower":
            for prima_data in emp_prima_data_empower:
                if prima_data.p_emp_id==emp.emp_id and prima_data.p_weekend_date==output_dt:
                    wbs22[f'O{counter2}'].value=prima_data.p_total_hours
                    check_emp=False
                    if int(wbs22[f'O{counter2}'].value)>45:
                            wbs22[f'R{counter2}'].value="Billable  Hour is more than 45."
                    break
            wbs22[f'P{counter2}'].value=0
            wbs22[f'Q{counter2}'].value=f"=P{counter2}=O{counter2}"
            if check_emp==True:
                wbs22[f'O{counter2}'].value=0
                wbs22[f'R{counter2}'].value="Emp data not found in PRIMA Report" 
            check_emp=True
            counter2+=1

def prima_grod():
    global emp_prima_data_grod
    emp_prima_data_grod=[]
    for row in range(5,wbs14.max_row+1):
        emp_id=wbs14.cell(row=row, column=20).value.strip()
        if emp_id:
            p_account_id=wbs14.cell(row=row, column=1).value.strip()
            p_work_item=wbs14.cell(row=row, column=3).value.strip()
            
            new_p_weekend_date=wbs14.cell(row=row, column=26).value
            new_p_activity_code=wbs14.cell(row=row, column=8).value.strip()
            if new_p_activity_code=="GN0035":
                cnb=int(wbs14.cell(row=row, column=40).value)
                new_p_total_hours=0
            else:
                new_p_total_hours=int(wbs14.cell(row=row, column=40).value)
                cnb=0
            match_emp=None
            for check_emp in emp_prima_data_grod:
                if check_emp.p_emp_id==emp_id:
                    if check_emp.p_weekend_date==new_p_weekend_date and check_emp.p_activity_code==new_p_activity_code:
                        match_emp=check_emp
                        break
            if match_emp:
                match_emp.update(new_p_weekend_date, new_p_activity_code, new_p_total_hours,cnb)
            else:
                emp_prima_data_grod.append(PRIMA_GROD(emp_id,p_account_id, p_work_item,new_p_weekend_date,new_p_activity_code,new_p_total_hours,cnb))
    
    check_emp=True
    counter3=2
    for emp in emp_data:
        if emp.lob=='RS-GROD':
            for prima_data in emp_prima_data_grod:
                if prima_data.p_emp_id==emp.emp_id and prima_data.p_weekend_date==output_dt:
                    wbs23[f'O{counter3}'].value=prima_data.p_account_id
                    wbs23[f'P{counter3}'].value=f"=O{counter3}=A{counter3}"
                    if prima_data.p_activity_code=="GN0035":
                        wbs23[f'T{counter3}'].value=prima_data.cnb
                    else:
                        wbs23[f'Q{counter3}'].value=prima_data.p_work_item
                        wbs23[f'R{counter3}'].value=f"=Q{counter3}=H{counter3}"
                        wbs23[f'S{counter3}'].value=prima_data.p_total_hours
                        wbs23[f'T{counter3}'].value=0
                    check_emp=False

                    
             
            try:
                if check_emp==True:
                    wbs23[f'S{counter3}'].value=0
                    wbs23[f'T{counter3}'].value=0
                    wbs23[f'W{counter3}'].value="Emp data not found in PRIMA Report"
                elif wbs23[f'S{counter3}'].value>45 or wbs23[f'T{counter3}'].value>5:
                    wbs23[f'W{counter3}'].value="Billable Hour is more than 45 or CNB more than 5."
                elif wbs23[f'J{counter3}'].value=="Ind" and wbs23[f'E{counter3}'].value!="DB retiree" and (int(wbs23[f'S{counter3}'].value)/int(wbs23[f'T{counter3}'].value)!=8) and wbs23[f'S{counter3}'].value<=45:
                        wbs23[f'W{counter3}'].value="Billable and CNB Hours are mismatched"
            except ZeroDivisionError:
                    wbs23[f'W{counter3}'].value="Billable Hour and CNB Hours are mismatched"
                
            check_emp=True
            counter3+=1

def ALT_report():
    epic_dt=[wbs15.cell(row=row, column=11).value for row in range(wbs15.min_row+1,wbs15.max_row)]
    if check_dt-timedelta(days=6) in epic_dt:
        counter3=2
        unpaid_checker=True
        check_emp=True
        for emp in emp_data:
            if  emp.lob=='RS-GROD':
                epic_unpaid=['Non-Billable time','Paid Time Off','ALT - Non-Billable Project']
                for row in range(wbs15.min_row+1,wbs15.max_row):
                    if wbs15.cell(row=row, column=1).value.upper()[:7]==wbs23[f'C{counter3}'].value and wbs15.cell(row=row, column=11).value==check_dt-timedelta(days=6):
                        epic=wbs15.cell(row=row, column=12).value
                        check_emp=False
                        if wbs15.cell(row=row, column=9).value.strip() in epic_unpaid:
                            unpaid_checker=False
                        break
                if check_emp==False and unpaid_checker==True:
                    wbs23[f'U{counter3}'].value=epic
                elif check_emp==False and unpaid_checker==False:
                    wbs23[f'U{counter3}'].value=epic
                    wbs23[f'W{counter3}'].value="Emp unbillable in EPIC report"
                elif check_emp==True:
                    wbs23[f'U{counter3}'].value=0
                
                if check_emp==True and wbs23[f'U{counter3}'].value!=wbs23[f'S{counter3}'].value and (wbs23[f'W{counter3}'].value=='' or wbs23[f'W{counter3}'].value==None):   
                    wbs23[f'W{counter3}'].value="Emp data not found in EPIC Report"
                elif check_emp==True and wbs23[f'U{counter3}'].value!=wbs23[f'S{counter3}'].value and  (wbs23[f'W{counter3}'].value!='' or wbs23[f'W{counter3}'].value!=None):
                    wbs23[f'W{counter3}'].value=str(wbs23[f'W{counter3}'].value)+", Billable and EPIC mismatched"

    
                wbs23[f'V{counter3}'].value=f"=U{counter3}=S{counter3}"

                counter3+=1
                unpaid_checker=True
                check_emp=True
    else:
        print('\033[31;1mWeekending Date not matched with ALT Timesheet\033[0m')
                #epic=EPIC_checker(wbs23[f'C{counter3}'].value,check_dt-timedelta(days=6))




Welcome()
input_dt=input('>>> Please Enter Weekending Date (DD-MM-YYYY): ')
date_checker(input_dt)

#date_checker('22-03-2024')

input_file_roster=input('>>> Please Enter the Roster file name: ')
wb11=load_workbook(f'{input_file_roster}.xlsx')
wbs11=wb11.active

output_report()
roster_report()

input_file_prima_ms=input('>>> Please Enter the PRIMA RS-MS file name: ')
if input_file_prima_ms=='skip' or input_file_prima_ms=='':
    print('\033[34;1mPRIMA RS-MS file is skipped\033[0m')
else:
    wb12=load_workbook(f'{input_file_prima_ms}.xlsx')
    wbs12=wb12.active
    prima_ms()

input_file_prima_emp=input('>>> Please Enter the PRIMA EMPOWER file name: ')
if input_file_prima_emp=='skip' or input_file_prima_emp=='':
    print('\033[34;1mPRIMA EMPOWER file is skipped\033[0m')
else:
    wb13=load_workbook(f'{input_file_prima_emp}.xlsx')
    wbs13=wb13.active
    prima_empower()
     
input_file_prima_grod=input('>>> Please Enter the PRIMA RS-GROD file name: ')
if input_file_prima_grod=='skip' or input_file_prima_grod=='':
    print('\033[34;1mPRIMA RS-GROD file is skipped\033[0m')
else:
    wb14=load_workbook(f'{input_file_prima_grod}.xlsx')
    wbs14=wb14.active
    prima_grod()

input_file_alt_timesheet=input('>>> Please Enter the ALT Timesheet file name: ')
if input_file_alt_timesheet=='skip' or input_file_alt_timesheet=='':
    print('\033[34;1mALT Timesheet file is skipped\033[0m')
else:
    df5=alt_file(f'{input_file_alt_timesheet}.xls')
    df5.to_excel('ALT_Timesheet_new.xlsx',index=False)
    wb15=load_workbook('ALT_Timesheet_new.xlsx')
    wbs15=wb15.active
    ALT_report()

'''
input_dt='29-03-2024'
date_checker(input_dt)
wb11=load_workbook('roster.xlsx')
wbs11=wb11.active
print('check1')
output_report()
roster_report()
print('check1')
wb12=load_workbook('PRIMA-MS.xlsx')
wbs12=wb12.active
prima_ms()
print('check2')
wb13=load_workbook('PRIMA-EMP.xlsx')
wbs13=wb13.active
prima_empower()
print('check3')
wb14=load_workbook('PRIMA-GROD.xlsx')
wbs14=wb14.active
prima_grod()
print('check4')
df5=pd.read_excel('ALT_Timesheet.xls')
df5.to_excel('ALT_Timesheet_new.xlsx',index=False)
wb15=load_workbook('ALT_Timesheet_new.xlsx')
wbs15=wb15.active
ALT_report()
print('check5')'''

# ============================================================================

         

        
        
#-------------------------------------------------------------------------------------------

# ------------------------- < Styling Sheets > ---------------------------------
background_color="dbcb7b"  #  background color
font_color="000000"        # Font color black 

header_fill=PatternFill(start_color=background_color, end_color=background_color, fill_type="solid")
header_font=Font(bold=True, color=Color(rgb=font_color))

border=Border(left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"))

#-------------------------------------------------------------------------------------------

#---------------------------------- < Header > -------------------------------------------
for cell in wbs21[1]:
    cell.fill=header_fill
    cell.font=header_font
for cell in wbs22[1]:
    cell.fill=header_fill
    cell.font=header_font
for cell in wbs23[1]:
    cell.fill=header_fill
    cell.font=header_font
#----------------------------------- < Border > -------------------------------------------

for row in range(wbs21.min_row, wbs21.max_row+1):
    for col in range(wbs21.min_column, wbs21.max_column+1):
        wbs21.cell(row=row,column=col).border=border
for row in range(wbs22.min_row, wbs22.max_row+1):
    for col in range(wbs22.min_column, wbs22.max_column+1):
        wbs22.cell(row=row,column=col).border=border
for row in range(wbs23.min_row, wbs23.max_row+1):
    for col in range(wbs23.min_column, wbs23.max_column+1):
        wbs23.cell(row=row,column=col).border=border

#-------------------------------------------------------------------------------------------

#-------------------------------- < SAVE REPORT > ------------------------------------------

def report_save():
    try:
        wb21.save(app_path+f'/BMS_Report_{input_dt_ot}.xlsx')
    except PermissionError:
        print(f'\033[31;1mBMS_Report_{input_dt_ot}.xlsx is opened. Please close the report before running the code. \033[0m')
        input("Press Enter to Continue: ")
        report_save()
#wb.save(f"{output_file}.xlsx")
report_save()

#-------------------------------------------------------------------------------------------

#-------------------------------- < OUTPUT PRINT > -----------------------------------------

time.sleep(2)
def delay_print(s):
    for c in s:
        sys.stdout.write(c)
        sys.stdout.flush()
        time.sleep(0.10)
delay_print('\n\n\033[32;1;1m>>> Successfully Generated Reportt!! ;)\n\n>>> Have a Good Day....Thank you!! \033[0m\n\n\n')

time.sleep(1)
for i in range(15,0,-1):
    print(f"\rWindow will close in 15 Seconds: \033[31;1;1m{i} \033[0m", end="", flush=True)
    time.sleep(1)

#-------------------------------------------------------------------------------------------


#===========================================================================================
#                                   < END OF PROGRAM >
#==========================================================================================